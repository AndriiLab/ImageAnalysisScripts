__author__ = 'AndriiLab'

""" Script calculates ratio between two channels ch1/ch2 from image series.
Next it compares two group of series (e.g. two proteins) and 
out xls file with results of comparison and graphs with plotted data"""

import cv2
import glob
from matplotlib import pyplot
import numpy
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from PIL import Image

def xlsFileWriter(filename, sheetname, dataname, x, y, yerr):
	"""stands for xls file writing:

		filename - name of your xls file
		sheetname - name of your xls sheetname
		dataname - name of grooped data (e.g. Protein 1 Experiment 1)
		x - list with X'es
		y - list with Y'es
		yerr - list with Y-errors

	Output: xls file with results"""
    global col
    #try to open existing workbook
    try:
        wb = load_workbook("%s.xlsx" %filename)
    except:
    #if workbook doesnt exist, creating it
        wb = Workbook()
    if sheetname in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheetname)
    else:
        ws = wb.create_sheet(title=sheetname)
    ws.cell(row=1, column=col, value="Folder name")
    ws.cell(row=1, column=col+1, value=dataname)
    #enter here label of X's values, Y's values
    ws.cell(row=2, column=col, value="X's name")
    ws.cell(row=2, column=col+1, value="Y's name")
    ws.cell(row=2, column=col+2, value="SD")
    for r, d1 in enumerate(x, 3):
        ws.cell(row=r, column=col, value=d1)
    for r, d2 in enumerate(y, 3):
        ws.cell(row=r, column=col+1, value=d2)
    for r, d3 in enumerate(yerr, 3):
        ws.cell(row=r, column=col+2, value=d3)
    wb.save("%s.xlsx" %filename)
    col+=4
    return True

def calcRatio(imgch1, imgch2, prot):
	"""calculates ratio between two images
		imgch1 - image with channel 1
		imgch2 - image with channel 2
		prot - switcher between two formulas for convetion from intensity to desired dimension
	Output: mean and stddev of ratio ch1/ch2"""
    global mean
    global stddev
    #converting channel 1 and 2 arrays to float data type
    ch1 = numpy.array(Image.open(imgch1)).astype(float)
    ch2 = numpy.array(Image.open(imgch2)).astype(float)

    #uncommit if you want to make lens normalization (removing dust, different intensity in regions etc)
    #you need to provide blank images made on your system in YOUR_PATH
    #imgch1corr = "YOUR_PATH"
    #imgch2corr = "YOUR_PATH"
    #ch1corr = numpy.array(Image.open(imgch1corr))/numpy.amax(numpy.array(Image.open(imgch1corr)))
    #ch2corr = numpy.array(Image.open(imgch2corr))/numpy.amax(numpy.array(Image.open(imgch2corr)))
    #ch1 = ch1/ch1corr
    #ch2 = ch2/ch2corr

    #filtering noise with Gaussian blur and Otsu thresholding
    mask = cv2.imread(imgch1, 0)
    blur = cv2.GaussianBlur(mask,(5,5),0)
    thresh, mask = cv2.threshold(blur,0,1,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    ch1 = ch1*mask
    ch2 = ch2*mask

    #removing 0es and max values (12bit=4095max)
    if 4095 in ch1:
        ch1[ch1==4095] = numpy.nan
    if 4095 in ch2:
        ch2[ch2==4095] = numpy.nan
    if 0 in ch1:
        ch1[ch1==0] = numpy.nan
    if 0 in ch2:
        ch2[ch2==0] = numpy.nan

    #calculating ratio 405/488 (ch2/ch1)
    ratio = ch2/ch1

    #uncommit if you want additionaly filter your data manually
    #mincutoff = 0.1
    #maxcutoff = 2.0
    #ratio[ratio < mincutoff] = numpy.nan
    #ratio[ratio > maxcutoff] = numpy.nan

    #for convetion from intensity to desired dimension enter formulae
    if prot =="sbb":
        ratio = ratio*1 #enter here formula for protein 1
    else:
        ratio = ratio*2	#enter here formula for protein 2

    #calculating mean and SD of the ratio
    ratio_wo_nan = numpy.ma.masked_array(ratio,numpy.isnan(ratio))
    mean.append(numpy.mean(ratio_wo_nan))
    stddev.append(numpy.std(ratio_wo_nan))
    return mean, stddev

def parseFolder(imgfolder):
	"""parses your folder for folders with images and sorts them as images for Protein 1
	and Protein 2. Next, calculates ratio between ch1 and ch2 for image series for this proteins.
	Plots data and saves xls file with results
		imgfolder - folder for parsing images
	Output: xls file with results and graph"""
    global mean
    global stddev
    global col
    datalistsbb = []
    datalistrd = []
    #retriving list of subdirs with data
    dirlist = next(os.walk(imgfolder))[1]
    #parsing for files with data
    for folder in dirlist:
        os.chdir(os.path.join(imgfolder+folder))
        imgch1 = []
        imgch2 = []
        mean = []
        stddev = []
        for file in glob.glob("C1*"): #you can change "C1" to desired channel 1 file name
            imgch1.append(file)
        for file in glob.glob("C[2-3]*"): #you can change "C[2-3]" to desired channel 2 file name
            imgch2.append(file)
        #calculating ratio for selected folder
        for i in imgch1:
            if "SbB" in folder: #you can change "SbB" to desired Protein 1 folder name
                calcRatio(i, imgch2[imgch1.index(i)],"sbb")
            elif "Rd" in folder: #you can change "Rd" to desired Protein 2 folder name
                calcRatio(i, imgch2[imgch1.index(i)],"rd")
            else:
                print "Protein not identified. Skipping folder."
        if "SbB" in folder:
            datalistsbb.append([folder, mean, stddev]) #writing data file for protein 1
        else:
            datalistrd.append([folder, mean, stddev]) #writing data file for protein 2

    #showing results as graph
    graprhheight = 2
    step = 1
    #set ymin and ymax
    ymin = 5.0
    ymax = 7.5
    col = 1
    os.chdir(imgfolder)
    for i in datalistrd:
        startpoint = float(i[0][-3:])
        pyplot.subplot(graprhheight,1,1), \
        pyplot.errorbar(numpy.linspace(startpoint, len(i[1])*step+startpoint-step,len(i[1])), i[1], yerr=i[2]),
        pyplot.ylim(ymin, ymax)
        pyplot.title("Protein 1")
        xlsFileWriter("results", "Protein 1", i[0],
                      numpy.linspace(startpoint, len(i[1])*step+startpoint-step,len(i[1])), i[1], i[2])

    col = 1
    for i in datalistsbb:
        startpoint = float(i[0][-3:])
        pyplot.subplot(graprhheight,1,2), \
        pyplot.errorbar(numpy.linspace(startpoint, len(i[1])*step+startpoint-step,len(i[1])), i[1], yerr=i[2]),
        pyplot.ylim(ymin, ymax)
        pyplot.title("Protein 2")
        xlsFileWriter("results", "Protein 1", i[0],
                      numpy.linspace(startpoint, len(i[1])*step+startpoint-step,len(i[1])), i[1], i[2])

    pyplot.xlabel("X's name")
    pyplot.ylabel("Y's name")
    pyplot.show()
    return True

if __name__ == "__main__":
    parseFolder("YOUR_PATH") #change YOUR_PATH with desired folder with subfolders with images

